#!/usr/bin/env python3
"""
sync-sanmar.py — SanMar catalog sync for the live pricing calculator.

Pipeline:
  1. SFTP connect to ftp.sanmar.com:2200 (SFTP, not FTPS).
  2. Download SanMar_SDL_N.csv (main catalog with image URLs) and
     sanmar_sale_items.txt (optional; pipe-delimited; has active sale prices).
  3. Parse the CSV: rows are per-SKU (style + color + size). Group by style.
  4. Apply filters.json rules (category/brand/style/discontinued/price).
  5. Merge custom-products.xlsx (non-SanMar items) if present.
  6. Write products.json matching the schema the calculator expects.

Environment variables (required):
  SANMAR_FTP_USER      = your SanMar customer number
  SANMAR_FTP_PASSWORD  = FTP password (NOT your sanmar.com password — they are different)
  SANMAR_FTP_HOST      = ftp.sanmar.com     (optional; defaults to this)
  SANMAR_FTP_PORT      = 2200               (optional; defaults to 2200)

Local use:
  python sync-sanmar.py                    # full run
  python sync-sanmar.py --dry-run          # parse but don't write
  python sync-sanmar.py --local PATH       # skip SFTP, use a local CSV
  python sync-sanmar.py --limit 50         # only process first 50 styles (for testing)
"""

import argparse
import csv
import io
import json
import os
import re
import sys
import zipfile
from datetime import datetime, timezone
from pathlib import Path

# Third-party deps (pinned in requirements.txt / GitHub Action)
import paramiko        # SFTP client
from openpyxl import load_workbook  # for custom-products.xlsx

# ---------------------------------------------------------------------------
# Config / constants
# ---------------------------------------------------------------------------
SANMAR_HOST_DEFAULT = "ftp.sanmar.com"
SANMAR_PORT_DEFAULT = 2200

SDL_FILENAMES = [
    # SanMar keeps the SDL file in the SanMarPDD folder. We try these paths
    # in order. Some accounts see it directly in the SFTP root.
    "SanMarPDD/SanMar_SDL_N.csv",
    "SanMar_SDL_N.csv",
    "SanMarPDD/Sanmar_SDL_N.csv",
    "Sanmar_SDL_N.csv",
]
SALE_ITEMS_FILENAMES = [
    "SanMarPDD/sanmar_sale_items.txt",
    "sanmar_sale_items.txt",
    "SanMarPDD/Sanmar_SaleItems.txt",
    "Sanmar_SaleItems.txt",
]

# Standard SanMar color hex codes. SanMar's feed gives color name but not hex;
# we lookup common names here. Anything not found falls back to #888888.
# You can extend this list over time.
COLOR_HEX = {
    "white": "#FFFFFF", "black": "#000000", "jet black": "#0A0A0A",
    "navy": "#1B2C55", "true royal": "#1F3FA3", "royal": "#1F3FA3",
    "red": "#B21F2D", "true red": "#B21F2D", "athletic red": "#C41E3A",
    "kelly green": "#008E5A", "forest green": "#2E5033", "dark green": "#2E5033",
    "yellow": "#FFE535", "gold": "#D4A52B", "athletic gold": "#D4A52B",
    "orange": "#F36F21", "safety orange": "#F36F21",
    "purple": "#603B8E", "athletic purple": "#603B8E",
    "pink": "#ED6EA7", "light pink": "#FBC1D6",
    "charcoal": "#3F4343", "dark heather": "#47484B",
    "athletic heather": "#A8ABAE", "ash": "#D4D2D0", "heather grey": "#B7B5B3",
    "sport grey": "#9EA5AE", "grey": "#8B8D8E",
    "brown": "#5C3A2E", "dark chocolate brown": "#3A2418",
    "sand": "#D4C4A8", "natural": "#E8DDC1", "khaki": "#A68B60",
    "maroon": "#66253F", "cardinal": "#8A1538",
    "light blue": "#A7C7E7", "carolina blue": "#8DB4D4", "sky": "#8DB4D4",
    "teal": "#1B6E7F", "mint": "#9FD8B8",
    "safety green": "#DAEF1E", "neon green": "#A4F713",
    "safety yellow": "#FFF33B",
    "heather navy": "#2B3A57", "heather red": "#A83C48",
    "heather royal": "#3E52A1",
    "coyote brown": "#72604A", "dark khaki": "#6F5B3E",
}

# Screen-print pricing matrix (per location, per color count, per qty break).
# Copied from the POC so the JSON can be published with default pricing embedded.
# You can override these in filters.json later if you want per-customer pricing.
SP_BREAKS = [24, 36, 50, 72, 144, 300, 500, 1000, 2500, 5000]
FRONT_SP = {
    24: {1: 2.85, 2: 3.75, 3: 4.50, 4: 5.25, 5: 6.00, 6: 6.75, 7: 7.50, 8: 8.25},
    36: {1: 2.40, 2: 3.15, 3: 3.85, 4: 4.55, 5: 5.20, 6: 5.85, 7: 6.50, 8: 7.15},
    50: {1: 2.10, 2: 2.75, 3: 3.40, 4: 4.00, 5: 4.55, 6: 5.10, 7: 5.65, 8: 6.20},
    72: {1: 1.85, 2: 2.45, 3: 3.00, 4: 3.50, 5: 4.00, 6: 4.50, 7: 5.00, 8: 5.50},
    144: {1: 1.55, 2: 2.00, 3: 2.45, 4: 2.90, 5: 3.30, 6: 3.70, 7: 4.10, 8: 4.50},
    300: {1: 1.30, 2: 1.70, 3: 2.05, 4: 2.40, 5: 2.75, 6: 3.05, 7: 3.35, 8: 3.65},
    500: {1: 1.15, 2: 1.50, 3: 1.80, 4: 2.10, 5: 2.40, 6: 2.65, 7: 2.90, 8: 3.15},
    1000: {1: 1.00, 2: 1.30, 3: 1.55, 4: 1.80, 5: 2.05, 6: 2.30, 7: 2.55, 8: 2.80},
    2500: {1: 0.90, 2: 1.15, 3: 1.40, 4: 1.60, 5: 1.85, 6: 2.05, 7: 2.25, 8: 2.45},
    5000: {1: 0.80, 2: 1.05, 3: 1.25, 4: 1.45, 5: 1.65, 6: 1.85, 7: 2.05, 8: 2.25},
}
SLEEVE_SP_ADDER = 1.50  # flat per-piece surcharge for sleeve location (screen print)
HT_BREAKS = [[10, 12.75], [20, 8.00], [50, 5.41], [100, 4.95], [200, 4.33],
             [300, 4.15], [500, 4.11], [1000, 4.08], [2500, 4.03], [5000, 4.00]]
HT_SLEEVE = 3.00
EM_BREAKS = [[12, 8.50], [24, 7.00], [50, 6.00], [100, 5.25], [200, 4.75], [500, 4.25]]

TURNAROUND = {
    "std":   {"days": 10, "surcharge_pct": 0},
    "rush":  {"days": 7,  "surcharge_pct": 15},
    "super": {"days": 5,  "surcharge_pct": 25},
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def log(msg):
    print(f"[sync-sanmar] {msg}", flush=True)


def size_sort_key(s):
    """Return a comparable key so sizes sort like a sane human expects."""
    order = ["XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL", "5XL", "6XL", "7XL",
             "XXL", "XXXL", "XXXXL"]
    s = s.strip().upper()
    if s in order:
        return (0, order.index(s))
    # youth sizes: YXS, YS, YM, YL, YXL
    if s.startswith("Y"):
        return (1, order.index(s[1:]) if s[1:] in order else 99)
    # numeric (like "4", "6", "8" for youth) or anything else — push to end
    return (2, s)


def guess_hex(color_name):
    """Best-effort hex lookup. Always returns a 7-char hex string."""
    key = (color_name or "").strip().lower()
    if key in COLOR_HEX:
        return COLOR_HEX[key]
    # Try the last word (e.g. "Heather Royal" -> "royal")
    parts = key.split()
    if parts and parts[-1] in COLOR_HEX:
        return COLOR_HEX[parts[-1]]
    return "#888888"


def normalize_size(s):
    """SanMar uses 'XX-Large', 'X-Small' etc. in some rows. Normalize to 'XL', 'XS'."""
    if not s:
        return ""
    s = s.strip().upper()
    mapping = {
        "X-SMALL": "XS", "SMALL": "S", "MEDIUM": "M", "LARGE": "L",
        "X-LARGE": "XL", "XX-LARGE": "2XL", "XXX-LARGE": "3XL",
        "XXXX-LARGE": "4XL", "XXXXX-LARGE": "5XL",
    }
    return mapping.get(s, s)


def parse_float(x):
    if x is None or x == "":
        return None
    try:
        return float(str(x).replace("$", "").replace(",", "").strip())
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# SFTP download
# ---------------------------------------------------------------------------
def download_sanmar_files(out_dir: Path):
    """SFTP in, download SDL + sale items. Returns (sdl_path, sale_path or None)."""
    user = os.environ.get("SANMAR_FTP_USER")
    pw = os.environ.get("SANMAR_FTP_PASSWORD")
    host = os.environ.get("SANMAR_FTP_HOST", SANMAR_HOST_DEFAULT)
    port = int(os.environ.get("SANMAR_FTP_PORT", SANMAR_PORT_DEFAULT))

    if not user or not pw:
        raise RuntimeError(
            "Missing SANMAR_FTP_USER / SANMAR_FTP_PASSWORD env vars. "
            "In GitHub Actions these come from repo Secrets."
        )

    log(f"Connecting to {host}:{port} as {user}…")
    transport = paramiko.Transport((host, port))
    # SanMar uses password auth; disable the GSS/kerberos probes that sometimes hang
    transport.connect(username=user, password=pw)
    sftp = paramiko.SFTPClient.from_transport(transport)

    def try_download(remote_candidates, local_name):
        """Try each candidate path; return local path on first success, or None."""
        for remote in remote_candidates:
            try:
                local_path = out_dir / local_name
                sftp.get(remote, str(local_path))
                log(f"  downloaded: {remote}  →  {local_path}")
                return local_path
            except FileNotFoundError:
                continue
            except IOError as e:
                # paramiko raises IOError for missing files on some servers
                if "No such file" in str(e) or "not found" in str(e).lower():
                    continue
                raise
        return None

    sdl_path = try_download(SDL_FILENAMES, "SanMar_SDL_N.csv")
    if not sdl_path:
        # List root so the user can see what's actually there
        log("Could not find SDL file at any expected path. Root directory listing:")
        for f in sftp.listdir("."):
            log(f"    {f}")
        try:
            for f in sftp.listdir("SanMarPDD"):
                log(f"    SanMarPDD/{f}")
        except IOError:
            pass
        raise RuntimeError("SDL file not found — see listing above")

    sale_path = try_download(SALE_ITEMS_FILENAMES, "sanmar_sale_items.txt")
    if not sale_path:
        log("  (sale-items file not found — continuing without sale prices)")

    sftp.close()
    transport.close()
    return sdl_path, sale_path


# ---------------------------------------------------------------------------
# CSV parsing
# ---------------------------------------------------------------------------
def parse_sale_items(sale_path: Path):
    """Parse sanmar_sale_items.txt (pipe-delimited) into a dict keyed by
    unique_key (INVENTORY_KEY + SIZE_INDEX)."""
    if not sale_path or not sale_path.exists():
        return {}
    sales = {}
    with open(sale_path, encoding="latin-1") as f:
        reader = csv.reader(f, delimiter="|", quotechar='"')
        header = next(reader, None)
        if not header:
            return {}
        # Column layout per the SanMar FTP Integration Guide v23.4:
        # 1 INVENTORY_KEY, 2 CATALOG_NO, 3 MILL, 4 MILL_STYLE_NO, 5 COLORCAT,
        # 6 COLOR, 7 SIZE, 8 DESCRIPTION, 9 EXT_DESC, 10 EA_SALE_PRICE,
        # 11 DZ_SALE_PRICE, 12 DZ_QTY, 13 CASE_SALE_PRICE, 14 CASE_QTY,
        # 15 CATALOG_PAGE, 16 PIECE_WEIGHT, 17 SIZE_TYPE, 18 SIZE_INDEX,
        # 19 STEVES_GROUP, 20 GTIN, 21 SALE_START_DATE, 22 SALE_END_DATE
        for row in reader:
            if len(row) < 22:
                continue
            inv_key = row[0].strip()
            size_idx = row[17].strip()
            unique_key = f"{inv_key}{size_idx}"
            sales[unique_key] = {
                "ea_sale_price": parse_float(row[9]),
                "case_sale_price": parse_float(row[12]),
                "sale_start_date": row[20].strip() or None,
                "sale_end_date": row[21].strip() or None,
            }
    log(f"  loaded {len(sales)} sale entries")
    return sales


def parse_sdl(sdl_path: Path, sales_by_key: dict, limit=None):
    """Parse SanMar_SDL_N.csv. Rows are keyed by UNIQUE_KEY (one per color+size).
    Group rows by style; within each style, group by color to collect
    per-color image URLs, sizes, and size upcharges."""
    styles = {}  # style_no -> product dict being built

    log(f"Reading {sdl_path}…")
    # SanMar's CSVs are quote-encapsulated, comma-delimited, latin-1 encoded
    with open(sdl_path, encoding="latin-1", newline="") as f:
        reader = csv.DictReader(f)
        # Normalize header keys: strip whitespace & uppercase. Field names from
        # the guide include e.g. STYLE# (with hash), BACK_MODEL _IMAGE_URL
        # (stray space), etc. We remap them.
        def norm(k):
            return re.sub(r"\s+", "", (k or "").upper()).rstrip("#")
        reader.fieldnames = [norm(h) for h in (reader.fieldnames or [])]

        row_count = 0
        for row in reader:
            row_count += 1
            row = {norm(k): (v.strip() if isinstance(v, str) else v) for k, v in row.items()}

            style = row.get("STYLE") or row.get("MILL_STYLE_NO")
            if not style:
                continue
            style = style.strip()

            color_name = row.get("COLOR_NAME", "").strip()
            if not color_name:
                continue

            size = normalize_size(row.get("SIZE", ""))

            case_price = parse_float(row.get("CASE_PRICE"))
            piece_price = parse_float(row.get("PIECE_PRICE"))
            msrp = parse_float(row.get("MSRP"))
            weight = parse_float(row.get("PIECE_WEIGHT"))

            # CATEGORY_NAME in SDL_N may contain semicolons separating main+sub
            cat_raw = row.get("CATEGORY_NAME", "")
            subcat = row.get("SUBCATEGORY_NAME", "")
            if ";" in cat_raw and not subcat:
                parts = [p.strip() for p in cat_raw.split(";") if p.strip()]
                category = parts[0] if parts else ""
                subcategory = "; ".join(parts[1:]) if len(parts) > 1 else ""
            else:
                category = cat_raw
                subcategory = subcat

            inv_key = row.get("INVENTORY_KEY", "").strip()
            size_idx = row.get("SIZE_INDEX", "").strip()
            unique_key = f"{inv_key}{size_idx}"
            sale = sales_by_key.get(unique_key, {})

            # Create or fetch the style entry
            if style not in styles:
                if limit is not None and len(styles) >= limit:
                    continue  # stop adding new styles once we hit the limit
                styles[style] = {
                    "style": style,
                    "name": row.get("PRODUCT_TITLE", "").strip(),
                    "brand": row.get("MILL", "").strip(),
                    "category": category,
                    "subcategory": subcategory,
                    "description": row.get("PRODUCT_DESCRIPTION", "").strip(),
                    "case_price": case_price,
                    "case_sale_price": None,
                    "sale_start_date": None,
                    "sale_end_date": None,
                    "msrp": msrp,
                    "piece_weight_lbs": weight,
                    "product_status": row.get("PRODUCT_STATUS", "").strip(),
                    "discontinued": (row.get("PRODUCT_STATUS", "").strip().lower() == "discontinued"),
                    "available_sizes": set(),
                    "size_upcharges": {},   # size -> upcharge vs. base case_price
                    "_base_case_price": case_price,  # reference for upcharge math
                    "_colors_by_name": {},
                }

            entry = styles[style]

            # Track smallest case_price as the "base" and anything above as an upcharge
            if case_price is not None:
                if entry["_base_case_price"] is None or case_price < entry["_base_case_price"]:
                    entry["_base_case_price"] = case_price
                    # If we just found a cheaper base, recompute existing upcharges
                    new_upcharges = {}
                    for s, up_cp in entry.get("_case_price_per_size", {}).items():
                        delta = round(up_cp - case_price, 2)
                        if delta > 0:
                            new_upcharges[s] = delta
                    entry["size_upcharges"] = new_upcharges
                entry.setdefault("_case_price_per_size", {})
                entry["_case_price_per_size"][size] = case_price
                delta = round(case_price - entry["_base_case_price"], 2)
                if delta > 0:
                    entry["size_upcharges"][size] = delta

            # Case sale price: take the lowest sale price across sizes for this style
            case_sale = sale.get("case_sale_price")
            if case_sale is not None:
                if entry["case_sale_price"] is None or case_sale < entry["case_sale_price"]:
                    entry["case_sale_price"] = case_sale
                    entry["sale_start_date"] = sale.get("sale_start_date")
                    entry["sale_end_date"] = sale.get("sale_end_date")

            if size:
                entry["available_sizes"].add(size)

            # Color
            if color_name not in entry["_colors_by_name"]:
                entry["_colors_by_name"][color_name] = {
                    "name": color_name,
                    "code": row.get("SANMAR_MAINFRAME_COLOR", "").strip()
                            or row.get("COLORCAT", "").strip()
                            or color_name[:4].upper(),
                    "hex": guess_hex(color_name),
                    "pms": row.get("PMS_COLOR", "").strip() or None,
                    "images": {
                        "front_model": row.get("FRONT_MODEL_IMAGE_URL", "").strip() or None,
                        "back_model": row.get("BACK_MODEL_IMAGE_URL", "").strip() or None,
                        "front_flat": row.get("FRONT_FLAT_IMAGE_URL", "").strip() or None,
                        "back_flat": row.get("BACK_FLAT_IMAGE_URL", "").strip() or None,
                        "side_model": None,  # SDL_N doesn't provide this; leave null
                    },
                }

    log(f"  parsed {row_count} SKU rows → {len(styles)} styles")
    return styles


def finalize_styles(styles):
    """Convert the in-progress dicts to the final products.json schema."""
    products = []
    for style, entry in styles.items():
        # Promote base case price (the cheapest across sizes) to case_price
        entry["case_price"] = entry["_base_case_price"]

        # Sort sizes in a sensible order
        sizes = sorted(entry["available_sizes"], key=size_sort_key)

        colors = list(entry["_colors_by_name"].values())

        product = {
            "style": entry["style"],
            "name": entry["name"],
            "brand": entry["brand"],
            "category": entry["category"],
            "subcategory": entry["subcategory"],
            "description": entry["description"],
            "case_price": entry["case_price"],
            "case_sale_price": entry["case_sale_price"],
            "sale_start_date": entry["sale_start_date"],
            "sale_end_date": entry["sale_end_date"],
            "msrp": entry["msrp"],
            "piece_weight_lbs": entry["piece_weight_lbs"],
            "product_status": entry["product_status"],
            "discontinued": entry["discontinued"],
            "available_sizes": sizes,
            "size_upcharges": entry["size_upcharges"],
            "colors": colors,
            "source": "sanmar",
        }
        products.append(product)

    # Sort alphabetically by style so diffs in git are readable
    products.sort(key=lambda p: p["style"])
    return products


# ---------------------------------------------------------------------------
# Filtering
# ---------------------------------------------------------------------------
def apply_filters(products, filters):
    def matches_any(value, patterns):
        if not patterns:
            return False
        v = (value or "").strip().lower()
        return any(v == p.strip().lower() for p in patterns)

    cat_wl = filters.get("category_whitelist") or []
    cat_bl = filters.get("category_blacklist") or []
    style_wl = filters.get("style_whitelist") or []
    style_bl = filters.get("style_blacklist") or []
    brand_wl = filters.get("brand_whitelist") or []
    brand_bl = filters.get("brand_blacklist") or []
    incl_disc = filters.get("include_discontinued", False)
    incl_cs = filters.get("include_coming_soon", False)
    min_p = filters.get("min_case_price")
    max_p = filters.get("max_case_price")
    require_image = filters.get("require_image", True)

    kept = []
    dropped_counts = {}

    def drop(reason):
        dropped_counts[reason] = dropped_counts.get(reason, 0) + 1

    for p in products:
        # Status
        status = (p.get("product_status") or "").strip().lower()
        if p.get("discontinued") and not incl_disc:
            drop("discontinued"); continue
        if status == "coming soon" and not incl_cs:
            drop("coming_soon"); continue

        # Style
        if style_wl and not matches_any(p["style"], style_wl):
            drop("style_not_whitelisted"); continue
        if matches_any(p["style"], style_bl):
            drop("style_blacklisted"); continue

        # Brand
        if brand_wl and not matches_any(p["brand"], brand_wl):
            drop("brand_not_whitelisted"); continue
        if matches_any(p["brand"], brand_bl):
            drop("brand_blacklisted"); continue

        # Category — SDL_N category can be 'Tee Shirts; 100% Cotton Tee' etc.
        # Match against the main category only (before the first ';')
        main_cat = (p.get("category") or "").split(";")[0].strip()
        if cat_wl and not matches_any(main_cat, cat_wl):
            drop("category_not_whitelisted"); continue
        if matches_any(main_cat, cat_bl):
            drop("category_blacklisted"); continue

        # Price
        cp = p.get("case_price")
        if cp is None:
            drop("no_case_price"); continue
        if min_p is not None and cp < min_p:
            drop("below_min_price"); continue
        if max_p is not None and cp > max_p:
            drop("above_max_price"); continue

        # Images — drop products where no color has any image
        if require_image:
            has_any_image = False
            for c in p.get("colors", []):
                imgs = c.get("images") or {}
                if any(imgs.get(k) for k in ("front_model", "back_model",
                                             "front_flat", "back_flat")):
                    has_any_image = True
                    break
            if not has_any_image:
                drop("no_image"); continue

        kept.append(p)

    log(f"Filter results: kept {len(kept)} / dropped {sum(dropped_counts.values())}")
    for reason, n in sorted(dropped_counts.items(), key=lambda x: -x[1]):
        log(f"    dropped for {reason}: {n}")
    return kept


# ---------------------------------------------------------------------------
# Custom products (xlsx)
# ---------------------------------------------------------------------------
def load_custom_products(xlsx_path: Path):
    """Read custom-products.xlsx and return a list of product dicts.

    Expected columns (first sheet, header row):
      style, name, brand, category, subcategory, description,
      case_price, case_sale_price, sale_start_date, sale_end_date,
      msrp, piece_weight_lbs, product_status, discontinued,
      available_sizes (comma-separated, e.g. "S,M,L,XL,2XL"),
      size_upcharges (JSON object string, e.g. '{"2XL": 2.00, "3XL": 3.00}'),
      colors (JSON array string — see template)

    Empty/missing rows are skipped.
    """
    if not xlsx_path.exists():
        return []

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    def col(row, name):
        if name in headers:
            return row[headers.index(name)]
        return None

    products = []
    for row in rows[1:]:
        style = col(row, "style")
        if not style:
            continue
        style = str(style).strip()

        sizes_raw = col(row, "available_sizes") or ""
        sizes = [s.strip() for s in str(sizes_raw).split(",") if s.strip()]

        upcharges_raw = col(row, "size_upcharges")
        if upcharges_raw:
            try:
                upcharges = json.loads(str(upcharges_raw))
            except json.JSONDecodeError:
                log(f"  warn: size_upcharges for {style} is not valid JSON — skipping that field")
                upcharges = {}
        else:
            upcharges = {}

        colors_raw = col(row, "colors")
        if colors_raw:
            try:
                colors = json.loads(str(colors_raw))
            except json.JSONDecodeError:
                log(f"  warn: colors for {style} is not valid JSON — using single default")
                colors = [{"name": "Default", "code": "DFLT", "hex": "#888888",
                           "images": {}}]
        else:
            colors = [{"name": "Default", "code": "DFLT", "hex": "#888888", "images": {}}]

        products.append({
            "style": style,
            "name": str(col(row, "name") or "").strip(),
            "brand": str(col(row, "brand") or "").strip(),
            "category": str(col(row, "category") or "").strip(),
            "subcategory": str(col(row, "subcategory") or "").strip(),
            "description": str(col(row, "description") or "").strip(),
            "case_price": parse_float(col(row, "case_price")),
            "case_sale_price": parse_float(col(row, "case_sale_price")),
            "sale_start_date": str(col(row, "sale_start_date") or "").strip() or None,
            "sale_end_date": str(col(row, "sale_end_date") or "").strip() or None,
            "msrp": parse_float(col(row, "msrp")),
            "piece_weight_lbs": parse_float(col(row, "piece_weight_lbs")),
            "product_status": str(col(row, "product_status") or "Regular").strip(),
            "discontinued": bool(col(row, "discontinued")),
            "available_sizes": sizes,
            "size_upcharges": upcharges,
            "colors": colors,
            "source": "custom",
        })
    log(f"  loaded {len(products)} custom products from {xlsx_path}")
    return products


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--filters", default="filters.json")
    ap.add_argument("--custom", default="custom-products.xlsx")
    ap.add_argument("--output", default="products.json")
    ap.add_argument("--local", help="Skip SFTP, read SDL CSV from this local path")
    ap.add_argument("--local-sale", help="Optional local path to sanmar_sale_items.txt")
    ap.add_argument("--limit", type=int, help="Process only the first N styles (for testing)")
    ap.add_argument("--dry-run", action="store_true", help="Don't write products.json")
    ap.add_argument("--workdir", default=".cache", help="Where to stash downloaded files")
    args = ap.parse_args()

    # Load filter rules
    filters_path = Path(args.filters)
    if not filters_path.exists():
        raise SystemExit(f"filters file not found: {filters_path}")
    with open(filters_path) as f:
        filters = json.load(f)
    filters = {k: v for k, v in filters.items() if not k.startswith("_")}  # strip comment keys
    log(f"Loaded filters: {sorted(filters.keys())}")

    # Get the SDL + sale-items files
    workdir = Path(args.workdir)
    workdir.mkdir(parents=True, exist_ok=True)

    if args.local:
        sdl_path = Path(args.local)
        sale_path = Path(args.local_sale) if args.local_sale else None
        log(f"Using local SDL file: {sdl_path}")
    else:
        sdl_path, sale_path = download_sanmar_files(workdir)

    # Parse
    sales = parse_sale_items(sale_path) if sale_path else {}
    styles = parse_sdl(sdl_path, sales, limit=args.limit)
    products = finalize_styles(styles)

    # Filter
    products = apply_filters(products, filters)

    # Merge custom
    if filters.get("merge_custom_products", True):
        custom_path = Path(args.custom)
        if custom_path.exists():
            custom = load_custom_products(custom_path)
            # Custom products override SanMar items with the same style number
            existing = {p["style"]: p for p in products}
            for c in custom:
                existing[c["style"]] = c
            products = sorted(existing.values(), key=lambda p: p["style"])
            log(f"After merging custom products: {len(products)} total")
        else:
            log(f"  no custom-products.xlsx found at {custom_path} — skipping merge")

    # Assemble final JSON
    output = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": f"SanMar SFTP ({SANMAR_HOST_DEFAULT}) + custom-products.xlsx",
        "markup_percent": filters.get("markup_percent", 50),
        "turnaround": TURNAROUND,
        "decoration_pricing": {
            "screen_print": {
                "breaks": SP_BREAKS,
                "front_back_per_color": FRONT_SP,
                "sleeve_per_color": SLEEVE_SP_ADDER,
                "_notes": "Per-piece price. Matrix is qty_break → {color_count → price}. Sleeve is a flat adder per color, not a separate matrix."
            },
            "heat_transfer": {
                "qty_breaks": HT_BREAKS,
                "sleeve_adder": HT_SLEEVE,
                "_notes": "Full color (no color selector). qty_breaks are [min_qty, per_piece_price]."
            },
            "embroidery": {
                "qty_breaks": EM_BREAKS,
                "_notes": "Per location. No color surcharge."
            },
        },
        "product_count": len(products),
        "products": products,
    }

    if args.dry_run:
        log(f"DRY RUN — would write {len(products)} products to {args.output}")
        return

    # Write output atomically
    tmp = Path(args.output + ".tmp")
    with open(tmp, "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    tmp.replace(args.output)
    log(f"Wrote {args.output} ({len(products)} products, "
        f"{tmp.stat().st_size if tmp.exists() else Path(args.output).stat().st_size} bytes)")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"FATAL: {e}")
        raise
